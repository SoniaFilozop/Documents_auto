import pymorphy2
from openpyxl import load_workbook

morph = pymorphy2.MorphAnalyzer()


class Sovetskiy:
    def __init__(self, ul):
        self.rayon = 'Советского'

    def get_rayon(self):
        return self.rayon


class Sovuch1S(Sovetskiy):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '1'
        self.ul = ['Генерала Перхоровича', 'Дорожная', 'Краснодонская', 'Юлюса Янониса']
        self.per = []
        self.pr = ['Поселок Придонской']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Sovuch2S(Sovetskiy):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '2'
        self.ul = ['Домостроителей', 'Пионеров',
                   'Космонавтов']
        self.per = ['Бригадный', 'Дозорный', 'Попутный', 'Производственный', 'Уютный', 'Чистый']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch3S(Sovetskiy):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '3'
        self.ul = ['Волнистая', 'Бородина', 'Пешестрелецкая',
                   'Холмистая',
                   'Семилукская']
        self.per = ['Автоматчиков', 'Газовый', 'Газопроводный']
        self.pr = ['Село Подклетное']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            if self.u in self.ul or self.u in self.per:
                return True
            return False


class Sovuch4S(Sovetskiy):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '4'
        self.ul = ['Молодогвардейцев', 'Олеко Дундича', 'Поселковая', 'Песчаная']
        self.per = ['Поселковый, Песчаный']
        self.pr = ['Поселок Шилово']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num, number):
        if num in self.pr:
            self.ul = ['Братьев Петровых', 'Вяземская', 'Генерала Платонова', 'Генерала Раевского', 'Генерала Халютина',
                       'Дубравная', 'Бульвар Воинской Славы', 'Генерала Шатилова', 'Долинная', 'Инженерная',
                       'Курчатова',
                       'Лазарева', 'Леонида Утесова',
                       'Малышевская', 'Междуреченская', 'Обнинская', 'Пойменная', 'Привольная', 'Розовая', 'Рубежная',
                       'Сусанинская', 'Теплоэнергетиков', 'Хохольская']
            self.per = ['Капитана Мягкова', 'Полковника Старкова']
            self.prochee = ['Фрунзе', 'в/ч 32891', 'в/ч 51025']
            if self.u in self.ul and number == 1:
                return True
            elif self.u in self.per and number == 2:
                return True
            elif self.u in self.prochee and number == 3:
                return True
            else:
                return False
        else:
            if self.u in self.ul or self.u in self.per or self.u:
                return True


class Sovuch5S(Sovetskiy):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '5'
        self.ul = ['Героев Сибиряков', 'Крымская', 'Жигулевская']
        self.per = []
        self.pr = ['Село Малышево', 'Поселок 1 Мая']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            if self.u in self.ul or self.u in self.per:
                return True
            else:
                return False


class Sovuch6S(Sovetskiy):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '6'
        self.ul = ['Богачева', 'Краснозвездная', 'Путиловская', 'Чеботарева', 'Патриотов', 'Б.Фестивальный']
        self.per = ['Дружный', 'Жигулевский']
        self.pr = ['Поселок Шилово', 'Шиловский массив', 'Пост ДПС ГИБДД на 214 км автодороги Курск - Саратов',
                   'Поселок Западный']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            if num == 'Поселок Шилово':
                self.ul = ['Буданцева', 'Ключникова',
                           'Полковника Богомолова', 'Шибилкина']
                if self.u in self.ul:
                    return True
            else:
                return True
        else:
            if self.u in self.ul:
                return True
            return False


class Sovuch7S(Sovetskiy):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '7'
        self.ul = [
            'Антокольского', 'Изобретателей', 'Писателя Маршака',
            'Конструкторов',
            'Магнитогорская', 'Малявина', 'Меркулова',
            'Нестерова', 'Шендрикова']
        self.per = ['Ани Скоробогатько', 'Антокольского', 'Архипова', 'Дьякова', 'И.Земнухова', 'Любы Шевцовой',
                    'Магнитогорский', 'Мухиной', 'Нестерова', 'Путиловский', 'Сергея Тюленина', 'Смирнова',
                    'Ульяны Громовой']
        self.pr = ['Поселок Тенистый', "Ветеран",
                   'Лазурит Тепличный',
                   'Придонье',
                   'Радуга',
                   'Тихий Дон',
                   'Тихий Дон-2',
                   'Тихий Дон-3',
                   'Тихий Дон-4']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


class Sovuch8S(Sovetskiy):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '8'
        self.ul = [
            'Любы Шевцовой', 'Пешеходная', 'Урожайная',
            'Депутатская', 'Космонавта Комарова', 'Олега Кошевого', 'Якира', 'Южно-Моравская']
        self.per = []
        self.pr = ['Поселок Подпольный']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


class Zheleznodorozhny:
    def __init__(self, ul):
        self.rayon = 'Железнодорожного'

    def get_rayon(self):
        return self.rayon


class Sovuch1Z(Zheleznodorozhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '1'
        self.ul = ['19 Стрелковой дивизии', '25 Января', 'Артёма', 'Барбюса', 'Береговая',
                   'Белгородская', 'Беломорская', 'Богатырская', 'Богучарская', 'Боровская', 'Братская', 'Вагонная',
                   'Венская', 'Восточная', 'Гоголевская', 'Деповская', 'Диспетчерская', 'Добрый Путь',
                   'Калининградская',
                   'Колхозный Путь', 'Коммунальная', 'Красина', 'Кубанская', 'Кузнечная', 'Куйбышева', 'К. Цеткин',
                   'Лермонтова', 'Листопадная', 'Литейная', 'Локомотивная', 'Луговая', 'Малиновского', 'Машинистов',
                   'Механическая', 'Молодежная', 'Моховая', 'Нариманова', 'Новая', 'Новый быт', 'Панфилова',
                   'Паровозная',
                   'Пионерская', 'Планетная', 'Пугачева', 'Р. Люксембург', 'Санникова', 'Северная', 'Сенная',
                   'Солидарности', 'Столярная', 'С. Халтурина', 'Салтыкова-Щедрина', 'Тепловозная', 'Токарная',
                   'Тургенева',
                   'Черноморская', 'Электровозная', 'Юности', 'Южная', 'Серафимовича']
        self.per = ['Артёма', 'Береговой', 'Гоголя', 'Деповский', 'Диспетчерский', 'Еловый', 'Инютинский',
                    'Колхозный', 'Кольцевой', 'Малярный', 'Молодежный', 'Нариманова', 'Пионерский', 'Прибрежный',
                    'Солнечный', 'Сретенский', 'Червонный', 'Серафимовича']
        self.pr = ['Проезд Богучарский', 'Тупик Деповский']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Sovuch2Z(Zheleznodorozhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '2'
        self.ul = ['Артамонова', 'Гаршина', 'Добролюбова', 'Запорожская']
        self.per = ['Добролюбова', 'Запорожский']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch3Z(Zheleznodorozhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '3'
        self.ul = ['Архитектора Троицкого', 'Богдана Хмельницкого', 'Изыскателей', 'Лизы Чайкиной', 'Суворова']
        self.per = ['Богдана Хмельницкого', 'Суворовский']
        self.pr = ['Проезд Рационализаторов', 'Боровской', 'Ближний Бор', 'Веселый', 'Дальний Бор', 'Кожевенный',
                   'Нижний', 'Маяки', 'Усманский', 'Пески', 'нп. Госзаповедник', 'Поселок Боровое', 'Поселок Водокачка',
                   'Поселок Краснолесный', 'Поселок Репное']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Sovuch4Z(Zheleznodorozhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '4'
        self.ul = ['Одинцова', 'Переверткина']
        self.per = ['Поселковый', 'Песчаный']
        self.pr = ['Поселок Сомово', 'Поселок Полыновка']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Sovuch5Z(Zheleznodorozhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '5'
        self.ul = ['Багратиона', 'Весенняя', 'Витрука', 'Гончаровская', 'Грибоедова', 'З. Космодемьянской', 'Землячки',
                   'Зубрилова', 'Жуковского', 'Измайловская', 'Комсомольская', 'Котовского',
                   'Краснофлотская', 'Красная Поляна', 'Куликовская',
                   'Набережная', 'Нахимовская', 'Остужева', 'Перекопская',
                   'Полтавская', 'Севастопольская', 'Спартаковская', 'Урывского',
                   'Филатова', 'Фурмановская', 'Грибоедова', 'Тракторная', 'Ушакова', 'Рокоссовского']
        self.per = ['Альпийский', 'Гранитный', 'Доблести', 'З. Космодемьянской', 'Измайловский', 'Колхозный',
                    'Корниловский', 'Морской', 'Международный', 'Нахимовский', 'Полтавский', 'Севастопольский',
                    'Флотский']
        self.pr = ['ПМС 81', 'Путевая машинная станция N 81']

        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Kominternovsky:
    def __init__(self, ul):
        self.rayon = 'Коминтерновского'

    def get_rayon(self):
        return self.rayon


class Sovuch1K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '1'
        self.ul = ['Академика Басова', 'Болховитинова', 'Веселая', 'Гайдара',
                   'Дачный поселок', 'Донская', 'Жемчужная', 'Загородная', 'Леваневского', 'Маршала Жукова', 'Павлова',
                   'Покровская', 'Просвещения', 'Победы', 'Республиканская', 'Ракитинская', 'Славы', 'Тульская',
                   'Чернышова']
        self.per = ['8 Марта', 'Александровский', 'Березки', 'Вокзальный', 'Зеленый', 'Исполкомовский', 'Ленина',
                    'Радиозаводской', 'Свердлова']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        return False


class Sovuch2K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '2'
        self.ul = ['5-й танковой армии', 'Актрисы Рощиной', 'Беспаловой',
                   'Историка Веселовского', 'Краеведа Зверевой', 'Маршала Катукова', 'Нагорная', 'Подгоренская',
                   'Радиозаводская', 'Танковая', 'Шолохова']
        self.per = ['Бабушкина', 'Веселый', 'Колодезный', 'Коминтерновский', 'Политехнический', 'Полковой', 'Полынный',
                    'Свободы']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch3K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '3'
        self.ul = ['Абызова', 'Айдаровская', 'Беговая', 'Алексея Геращенко', 'Дмитрия Горина', 'Дружинников',
                   'Ипподромная', 'Карпинского', 'Красных Зорь', 'Независимости', 'Очаковская', 'Преображенская',
                   'Троепольского', 'Чкалова', 'Фруктовая', 'Электросигнальная']
        self.per = ['Калинина', 'Пугачева', 'Самойловский', 'Ученический']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch4K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '4'
        self.ul = ['Победы Б-Р', 'Архитектора Быховского', 'Архитектурная', 'Березовская', 'Багрицкого',
                   'Екатерины Зеленко', 'Малаховского', 'Остроухова', 'Серафима Саровского', 'Товарищеская']
        self.per = ['Дачный', 'Камышовый', 'Партизанский', 'Пейзажный', 'Правды', 'Студеный']
        self.pr = ['Бульвар Победы', 'Проезд Ясный']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Sovuch5K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '5'
        self.ul = ['Вокзальная', 'Гжельская', 'Головина', 'Дружеская',
                   'Ивана Туркенича', 'Курортная', 'Керамическая',
                   'Солнечная',
                   'Урицкого', 'Утренняя', 'Школьная']
        self.per = ['Анненский', 'Валуйский', 'Вольный', 'Грибановский', 'Дружеский', 'Красно-Лимановский', 'Станичный']
        self.pr = ['Территория Аэропорт']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Sovuch6K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '6'
        self.ul = ['Ведугская', 'Героев Бреста', 'Лебедянская', 'Миронова', 'Скляевой', 'Связистов', 'Сочинская']
        self.per = ['Елочный', 'Осинки', 'Прохоровский', 'Республиканский', 'Федеративный']
        self.pr = ['Жилой массия Ясенки']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Sovuch7K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '7'
        self.ul = ['45 Стрел.Дивизии', 'Елецкая',
                   'Историка Костомарова' 'Кораблинова', 'Криворучко', 'Терешковой', 'Художника Лихачева',
                   'Художника Пономарева']
        self.per = ['Автогенный', 'Верещагина', 'Гагарина', 'Дроздовый', 'Коллективный', 'Лискинский',
                    'Приветливый',
                    'Путейский Городок']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch8K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '8'
        self.ul = ['Труда',
                   '1 Мая',
                   'Верещагина', 'Вольная', 'Десантников', 'Землянская', 'Задонская', 'Композитора Ставонина',
                   'Минеров', 'Мордасовой', 'Новгородская', 'Овражная', 'Питомник', 'Правды', 'Свердлова',
                   'Социалистическая', 'Шукшина']
        self.per = ['Заветный', 'Загорский', 'Закатный', 'Здоровья', 'Клинический', 'Новоселов', 'Осетровский',
                    'Спокойный', 'Станкостроительный', 'Ракетный']
        self.pr = ['Жилой массив Задонье', 'Площадь Советов']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


class Sovuch9K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '9'
        self.ul = ['303 Стрелковой Дивизии', '60 Армии', 'Антонова-Овсеенко', 'Апраксина', 'Брянская', 'Ватутина',
                   'Генерала Ефремова', 'Княжеская', 'Лидии Рябцевой',

                   'Поэта Прасолова', 'Рязанская', 'Строителей', 'Тамбовская', 'Федора Сушкова', 'Фрегатная']
        self.per = ['Настасьинский', 'Октябрьский', 'Подклетенский', 'Рамонский', 'Солнечный', 'Школьный', 'Ямской']
        self.pr = ['Проезд Брянский', 'Жилой массив Хвойный']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


class Sovuch10K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '10'
        self.ul = ['Багряная', 'Братьев Мариных', 'Веневитинская', 'Любимая',
                   'Окружная', 'Строительная', 'Торпедо', 'Текстильщиков', 'Чудесная']
        self.per = ['Дальний', 'Космонавтов', 'Надежды', 'Озерный', 'Хреновской']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        return False


class Sovuch11K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '11'
        self.ul = ['65 лет Победы', 'Академика Королева', 'Варейкиса', 'Еремеева',
                   'Киевская', 'Крайняя', 'Курская', 'Маршала Голикова', 'Народная', 'Подклетенская', 'Церковная']
        self.per = ['Мечникова', 'Славы', 'Сливовый', 'Транспортный', 'Чесменский']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        return False


class Sovuch12K(Kominternovsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '12'
        self.ul = ['60 лет ВЛКСМ', 'Лесхозная',
                   'Новый поселок', 'Олифиренко']
        self.per = ['Ботанический', 'Снежный']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        return False


class Levoberezhny:
    def __init__(self, ul):
        self.rayon = 'Левобережного'

    def get_rayon(self):
        return self.rayon


class Sovuch1Lv(Levoberezhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '1'
        self.ul = ['Азовская', 'Волжская', 'Донецкая', 'Костромская', 'Летчика Небольсина', 'Новороссийская',
                   'Просторная', 'Рижская', 'Ржевская', 'Чебышева', 'Шубина']
        self.per = ['Новороссийский', 'Прохладный', 'Ростовский', 'Цимлянский']
        self.pr = ['Воронежсинтезкаучук ', 'Поселок Алексеевка', 'Южная подстанция']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Sovuch2Lv(Levoberezhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '2'
        self.ul = ['Глинки', 'Дубровина', 'Корольковой']
        self.per = []
        self.pr = ['Совхоз Никольский', 'Поселок Буденный', 'Поселок Масловка', 'Поселок Подклетный', 'Село Никольское',
                   'Семилукские Выселки', 'Село Таврово', 'Воронежский гидроузел']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


class Sovuch3Lv(Levoberezhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '3'
        self.ul = ['Балашовская', 'Барнаульская', 'Варшавская',
                   'Венецианова', 'Волгодонская', 'Волкова', 'Вологодская', 'Воронихина', 'Вьетнамская', 'Дружбы',
                   'Заслонова', 'Ковалевской', 'Корейская', 'Кронштадтская', 'Ленская', 'Макаренко', 'Мечникова',
                   'Новикова', 'Нововоронежская', 'Пекинская', 'Печерская',
                   'Путилина', 'Саврасова', 'Третьякова', 'Цимлянская', 'Читинская',
                   'Шахтинская', 'Шинников']
        self.per = ['Баженова', 'Балашовский', 'Заслонова', 'Ковыльный', 'Химический', 'Шубина']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch4Lv(Levoberezhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '4'
        self.ul = ['Беляевой', 'Героев Стратосферы', 'Кулибина', 'Лебедева', 'Левитана',
                   'Лихачева', 'Менделеева', 'Меркулова', 'Полины Осипенко',
                   'Осоавиахимовская', 'Сеченова', 'Танеева', 'Тельмана', 'Ярославская']
        self.per = ['Русанова', 'Сеченова', 'Тельмана']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch5Lv(Levoberezhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '5'
        self.ul = ['Арзамасская',
                   'Круглая', 'МОПРА', 'Нижняя', 'Порт-Артурская', 'Спортивная набережная', 'Тихая', 'Щорса']
        self.per = ['Арзамасский', 'Гвардейский', 'Заречный', 'Кавказский', 'Казанский', 'Мостостроителей', 'Ольховый',
                    'Парашютистов', 'Планерный', 'Сквозной', 'Сторожевой']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch6Lv(Levoberezhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '6'
        self.ul = ['Брусилова',
                   'Веры Засулич', 'Красногвардейская', 'Огарева',
                   'Озерная', 'Саперная', 'Серова']
        self.per = ['Бобровский', 'Гражданский', 'Крамского', 'Международный', 'Набережный', 'Ольховатский',
                    'Поленова' 'Репина', 'Усманский', 'Хоперский']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch7Lv(Levoberezhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '7'
        self.ul = ['17 Сентября', 'Витебская', 'Волгоградская', 'Героев Хасана', 'Гомельская', 'Дальневосточная',
                   'Ивановская', 'Калачеевская', 'Клинская', 'Новохоперская',
                   'Обручева', 'Окружная', 'Тверская', 'Туполева', 'Уточкина']
        self.per = ['Бродского', 'Брусилова', 'Витебский', 'Волочаевский', 'Димитрова', 'Ивановский',
                    'Путевой', 'Россошанский', 'Таловский',
                    'Томский', 'Уточкина', 'Чаплыгина', 'Челябинский']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch8Lv(Levoberezhny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '8'
        self.ul = ['Адмирала Чурсина', 'Айвазовского', 'Алданская', 'Базовая', 'Баррикадная', 'Владимира Высоцкого',
                   'Гаражный тупик', 'Ермака', 'Ильюшина', 'Иркутская', 'Красного Октября', 'Омская', 'Писарева',
                   'Ползунова', 'Славянская', 'Уральская', 'Циолковского', 'Шидловского', 'Черепанова', 'Черниговская']
        self.per = ['Монтажный', 'Отличников']
        self.pr = ['Монтажный', 'Придача', 'СМП N 808', 'СМП N 863', 'Озерки']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


class Leninsky:
    def __init__(self, ul):
        self.rayon = 'Ленинского'

    def get_rayon(self):
        return self.rayon


class Sovuch1Ln(Leninsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '1'
        self.ul = [
            '40-Летия Октября', 'Бакунина',
            'Девицкий выезд',

            'Пограничная', 'Промышленная',
            'Профсоюзная', 'Радищева']
        self.per = ['Бакунинский', 'Кабельный', 'Мало - Московский', 'Слесарный', 'Специалистов', 'Электронный']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        return False


class Sovuch2Ln(Leninsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '2'
        self.ul = [
            'Белинского', 'Бестужева',
            'Большая Стрелецкая',
            'Гора Металлистов', 'Кирова',
            'Коперника',
            'Красных Партизан', 'Крупской',
            'Льва Толстого', 'Мало-Стрелецкая', 'Нарвская',
            'Петра Сазонова',
            'Платонова', 'Ремесленная Гора',
            'Севастьяновский съезд', 'Серго',
            'Фрунзе',
            'Шевченко']
        self.per = ['Белинского', 'Бестужева', 'Свечной', 'Муравьева', 'Корабельный', 'Красных Партизан',
                    'Печатников',
                    'Рабкоровский', 'Сазонова', 'Севастьяновский']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch3Ln(Leninsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '3'
        self.ul = ['20-Летия Октября',
                   'Артековская', 'Бетховена',
                   'Волоколамская',
                   'Державина',
                   'Донбасская', 'Карамзина',
                   'Красная Горка', 'Красноярская',
                   'Кубанская Гора', 'Лескова', 'Майкова', 'Марата', 'Народных Ополченцев',
                   'Песчаная Гора', 'Свободная', 'Скрибиса', 'Саратовская', 'Станкевича',
                   'Челюскинцев', 'Чернышевский бугор', 'Черновицкая']
        self.per = ['Алтайский', 'Ангарский', 'Веры Фигнер', 'Городской', 'Жасминный', 'Заозерный',
                    'Карпатский', 'Лескова', 'Марата', 'Мурманский', 'Нансена', 'Невский',
                    'Поворотный', 'Полярный', 'Свободный', 'Снайперский', 'Цветной', 'Челюскинцев']
        self.pr = ['Маяковского', 'Балтийский']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


class Sovuch4Ln(Leninsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '4'
        self.ul = ['100 Стрелковой Дивизии', 'Аксакова', 'Возрождения', 'Верхняя', 'Грамши', 'Дзотова',
                   'Звенигородская', 'Курганская',
                   'Лишенко',
                   'Маленькая', 'Маяковского', 'Направника', 'Общественная',
                   'Профессора Харина', 'Римского-Корсакова', 'Серпуховская',
                   'Солодовникова', 'Средняя', 'Степана Панкова', 'Тушинская', 'Ульяновская',
                   'Уфимская', 'Уфимский тупик', 'Чижовская']
        self.per = ['Аксакова', 'Байкальский', 'Виноградова', 'Вишневый', 'Возрождения', 'Грамши', 'Даргомыжского',
                    'Керченский', 'Кирпичный', 'Коллективизации', 'Ленский', 'Можайский', 'Симферопольский',
                    'Сивашский', 'Солодовникова', 'Чапаева', 'Черкасский']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch5Ln(Leninsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '5'
        self.ul = ['Артиллерийская',
                   'Выборгская',
                   'Демьянова', 'Енисеевская', 'Енисеевский бугор',
                   'Злобина', 'Конноармейская',
                   'Конно-Стрелецкая', 'Кропоткина',
                   'Молдавская', 'Маршала Неделина',
                   'Малокрасноармейская', 'Пестеля',
                   'Петровская набережная',
                   'Пушкарская',
                   'Сергея Лазо', 'Танкистов',
                   'Черняховского', 'Щербакова']
        self.per = ['Амурский', 'Благодатный', 'Бессарабский', 'Карельский', 'Казарменный', 'Краснознамённый',
                    'Кедровый', 'Кишиневский', 'Конноармейский',
                    'Люлина', 'Конно-Стрелецкий', 'Минина', 'Молдавский',
                    'Малокрасноармейский', 'Ново-Слободский', 'Новый', 'Ореховый', 'Пушкарский', 'Пестеля',
                    'Пожарского', 'Пролетарский', 'Раздольный', 'Танкистов', 'Штурмовой']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch6Ln(Leninsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '6'
        self.ul = ['121 Стрелковой Дивизии',
                   'Андреева',
                   'Волнухина', 'Врубеля', 'Газетчиков',
                   'Голубиный тупик',
                   'Гремяченская', 'Журналистов', 'Калужская', 'Камская',
                   'Ладожская', 'Ладожский проезд',
                   'Новаторов',
                   'Новогодняя', 'Орловская',
                   'Острогожский проезд',
                   'Охотничий тупик',
                   'Праздничная', 'проезд Яблочный', 'Рабочий тупик', 'Рыбачий тупик',
                   'Соколиный тупик', 'Туркменский проезд', 'Ударная',
                   'Херсонская', 'Маршала Шапошникова', 'Шиловская',
                   'Энтузиастов']
        self.per = ['30-Летия Октября', 'Астраханский', 'Госпитальный', 'Грибной', 'Журналистов', 'Короткий',
                    'Кривошеина', 'Кленовый', 'Каштановый', 'Корпусный', 'Кривоколенный', 'Камский', 'Летний',
                    'Лиственный', 'Летчиков', 'Лучевой', 'Малый', 'Одесский', 'Орловский',
                    'Острогожский', 'Прямой', 'Рябиновый', 'Стрелковый', 'Соколиный',
                    'Хвойный', 'Херсонский', 'Ялтинский', 'Яблочный']
        self.pr = ['30-Летия Октября', 'Иртышский', 'Летний', 'Ялтинский', 'Малый']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


class Sovuch7Ln(Leninsky):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '7'
        self.ul = ['Астраханская', 'Базарная Гора', 'Бархатный Бугор', 'Броневая', 'Есенина',
                   'Красноармейская',
                   'Некрасова',
                   'Одесская',
                   '30-Летия Октября']
        self.per = ['Бакинский', 'Балтийский', 'Днепровский', 'K.Булавина', 'Красноармейский', 'Передовой', 'Старинный']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Сentralny:
    def __init__(self, ul):
        self.rayon = 'Центрального'

    def get_rayon(self):
        return self.rayon


class Sovuch1Cn(Сentralny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '1'
        self.ul = ['Академика Конопатова', 'Берёзовая Роща', 'Бунакова', 'Бурденко', 'Гастелло', 'Дарвина', 'Докучаева',
                   'Дуговая', 'Железноводская', 'Казакова', 'Калинина', 'Келлера', 'Кисловодская',
                   'Кисловодский проезд', 'Красной Работницы', 'Красовского', 'Ленина', 'Лесной поселок',
                   'Лобачевского', 'Ломоносова', 'Максима Горького набережная', 'Мичурина', 'Морозова',
                   'Октябрьской Революции', 'Поспеева', 'Прянишникова',
                   'Пятигорская', 'Рабочего класса',
                   'Рылеева', 'Серебрякова', 'Советская', 'Студенческая дом N 17', 'Тимирязева', 'Ушинского',
                   'Фронтовая', 'Цветущая', 'Швейников']
        self.per = ['Бауманский', 'Дуговой', 'Ессентуки', 'Курортный', 'Лесной', 'Лесокультурный', 'Лечебный',
                    'Опытный', 'Проходной', 'Советский', 'Спортивный', 'Учебный']
        self.pr = ['Санаторий им. Горького', 'Лесная поляна', 'Ботанический сад', 'Учебный кордон', 'Рыбачье', 'Ветряк',
                   'Олимпик']
        super().__init__(ul)

    def get(self):
        return self.sovuch

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get_pr(self, num):
        if num in self.pr:
            return True
        else:
            return False


class Sovuch2Cn(Сentralny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '2'
        self.ul = ['Площадь Университетская', 'Б-Р Олимпийский', '20-Летия ВЛКСМ', '25 Октября',
                   'Авиационная', 'Алексеевского', 'Белинского',
                   'Бехтерева',
                   'Василия Пескова',
                   'Герцена', 'Декабристов',
                   'Детей', 'Дзержинского', 'Загоровского', 'Кардашова', 'Кости Стрелюка', 'Левая Суконовка',
                   'Петра Алексеева', 'Первомайская',
                   'Правая Суконовка', 'Помяловского',
                   'Сакко и Ванцетти', 'Театральная', 'Чернышевского']
        self.per = ['Бехтерева', 'Горный', 'Детский', 'Мордовцева', 'Рабфаковский', 'Скорняжный',
                    'Старомосковский',
                    'Тимуровцев']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch3Cn(Сentralny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '3'
        self.ul = ['Академика Першина', 'Бовкуна', 'Бунина', 'Вавилова', 'Героев Красной Армии', 'Горького',
                   'Кавалерийская',
                   'Коммунаров', 'Короленко', 'Красненькая', 'Крестьянская', 'Крутая', 'Марии Копыловой', 'Линейная',
                   'Логвиновская', 'Луначарского', 'Пятницкого', 'Сиреневая',
                   'Софьи Перовской', 'Таранченко']
        self.per = ['Академика Гмелина']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        return False


class Sovuch4Cn(Сentralny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '4'
        self.ul = ['Генерала Черняховского',
                   'Городовикова', 'Желябова',
                   'Козо-Полянского',
                   'Мира', 'Нарядная',
                   'Обороны Революции', 'Полевая',
                   'Ср.Московская', 'Студенческая',
                   'Феоктистова', 'Чайковского', 'Юных Натуралистов',
                   'Коммунистической молодежи.']
        self.per = ['Героев Революции', 'Индустриальный', 'Купянский', 'Мельничный', 'Оранжерейный', 'Отрадный',
                    'Разведчиков', 'Учительский']
        self.pr = ['Жилой массив Олимпийский']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


class Sovuch5Cn(Сentralny):
    def __init__(self, ul):
        self.u = ul
        self.sovuch = '5'
        self.ul = ['Арсенальная', 'Батуринская',
                   'Большая Манежная', 'Большая Чернавская', 'Бучкури', 'Вайцеховского',
                   '8 Марта', 'Героев Революции', 'Героев Труда', 'Демократии', 'Дзиньковского', 'Достоевского',
                   'Дурова', 'Жилина', 'Замкина', 'Каляева',
                   'Комиссаржевской',
                   'Летчика Филипова', 'Малая Манежная', 'Мало-Смоленская', 'Мало-Терновая', 'набережная Массалитинова',
                   'Ольминского', 'Освобождения Труда', 'Пролетарская',
                   'Пролетарской Диктатуры', 'Смоленская', 'Ст.Разина', 'Трудовая', 'Целинная', 'Цюрупы', 'Эртеля']
        self.per = ['Пряничный', 'Солдатский', 'Фабричный', 'Целинный', 'Эртеля']
        self.pr = ['Рабочий городок']
        super().__init__(ul)

    def get_sovuch(self, number):
        if self.u in self.ul and number == 1:
            return True
        elif self.u in self.per and number == 2:
            return True
        else:
            return False

    def get(self):
        return self.sovuch

    def get_pr(self, num):
        if num in self.pr:
            return self.sovuch
        else:
            return False


def main():
    final = []
    va = []
    file_name = input('Введите название файла с таблицой с указанием формата файла. Пример: file.xlsx\n')
    wb2 = load_workbook(file_name)
    ws = wb2.active
    schet = -1
    for row in ws.values:
        for value in row:
            if type(value) is str:
                if value[:2] == 'Г.' or value[:2] == 'г.':
                    value = value[12:].split(',')
                    for i in value:
                        schet += 1
                        if i[0] == ' ':
                            value[schet] = i[1:]
                    schet = -1
                    va.append(value)
    for i in va:
        kolvo = 0
        if i[0][:3] == 'Ул.' or i[0][-2:] == 'Ул' or i[0][:4] == 'Пр-Т' or i[0][-3:] == 'Б-Р' or \
                i[0][:3] == 'ул.' or i[0][:4] == 'пр-т':
            number = 1
        elif i[0][:4] == 'Пер.' or i[0][-3:] == 'Пер' or i[0][:4] == 'пер.':
            number = 2
        else:
            number = 3
        if number == 3:
            v = '1'
        else:
            v = ''
        if v == '1':
            num = i[0]
            rayon = Sovetskiy(num)
            ul = ''
            if num == 'Поселок Шилово':
                if i[1][:3] == 'Ул.' or i[0][-2:] == 'Ул' or i[0][:4] == 'Пр-Т':
                    number = 1
                    ul = i[1][4:]
                elif i[1][:4] == 'Пер.':
                    ul = i[1][5:]
                    number = 2
                else:
                    ul = i[1]
                    number = 3
            else:
                number = 0
            s1 = Sovuch1S(ul)
            s2 = Sovuch2S(ul)
            s3 = Sovuch3S(ul)
            s4 = Sovuch4S(ul)
            s5 = Sovuch5S(ul)
            s6 = Sovuch6S(ul)
            s7 = Sovuch7S(ul)
            s8 = Sovuch8S(ul)
            if ul == 'Острогожская' and (int((i[2].split())[1]) / 2 == 0) and (int((i[2].split())[1]) >= 146):
                num = s6.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            else:
                if s1.get_pr(num):
                    num = s1.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s2.get_pr(num):
                    num = s2.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s3.get_pr(num):
                    num = s3.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s4.get_pr(num, number):
                    num = s4.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s5.get_pr(num):
                    num = s5.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s6.get_pr(num):
                    num = s6.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s7.get_pr(num):
                    num = s7.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s8.get_pr(num):
                    num = s8.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                s1 = Sovuch1Z(ul)
                s2 = Sovuch2Z(ul)
                s3 = Sovuch3Z(ul)
                s4 = Sovuch4Z(ul)
                s5 = Sovuch5Z(ul)
                if s1.get_pr(num):
                    num = s1.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s2.get_pr(num):
                    num = s2.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s3.get_pr(num):
                    num = s3.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s4.get_pr(num):
                    num = s4.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s5.get_pr(num):
                    num = s5.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                rayon = Kominternovsky(num)
                ul = ''
                s1 = Sovuch1K(ul)
                s2 = Sovuch2K(ul)
                s3 = Sovuch3K(ul)
                s4 = Sovuch4K(ul)
                s5 = Sovuch5K(ul)
                s6 = Sovuch6K(ul)
                s7 = Sovuch7K(ul)
                s8 = Sovuch8K(ul)
                s9 = Sovuch9K(ul)
                s10 = Sovuch10K(ul)
                s11 = Sovuch11K(ul)
                s12 = Sovuch12K(ul)
                if s1.get_pr(num):
                    num = s1.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s2.get_pr(num):
                    num = s2.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s3.get_pr(num):
                    num = s3.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s4.get_pr(num):
                    num = s4.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s5.get_pr(num):
                    num = s5.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s6.get_pr(num):
                    num = s6.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s7.get_pr(num):
                    num = s7.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s8.get_pr(num):
                    num = s8.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s9.get_pr(num):
                    num = s9.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s10.get_pr(num):
                    num = s10.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s11.get_pr(num):
                    num = s11.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s12.get_pr(num):
                    num = s12.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                rayon = Levoberezhny(num)
                ul = ''
                s1 = Sovuch1Lv(ul)
                s2 = Sovuch2Lv(ul)
                s3 = Sovuch3Lv(ul)
                s4 = Sovuch4Lv(ul)
                s5 = Sovuch5Lv(ul)
                s6 = Sovuch6Lv(ul)
                s7 = Sovuch7Lv(ul)
                s8 = Sovuch8Lv(ul)
                if s1.get_pr(num):
                    num = s1.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s2.get_pr(num):
                    num = s2.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s3.get_pr(num):
                    num = s3.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s4.get_pr(num):
                    num = s4.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s5.get_pr(num):
                    num = s5.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s6.get_pr(num):
                    num = s6.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s7.get_pr(num):
                    num = s7.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s8.get_pr(num):
                    num = s8.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                rayon = Leninsky(num)
                ul = ''
                s1 = Sovuch1Ln(ul)
                s2 = Sovuch2Ln(ul)
                s3 = Sovuch3Ln(ul)
                s4 = Sovuch4Ln(ul)
                s5 = Sovuch5Ln(ul)
                s6 = Sovuch6Ln(ul)
                s7 = Sovuch7Ln(ul)
                if s1.get_pr(num):
                    num = s1.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s2.get_pr(num):
                    num = s2.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s3.get_pr(num):
                    num = s3.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s4.get_pr(num):
                    num = s4.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s5.get_pr(num):
                    num = s5.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s6.get_pr(num):
                    num = s6.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s7.get_pr(num):
                    num = s7.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                rayon = Сentralny(num)
                ul = ''
                s1 = Sovuch1Cn(ul)
                s2 = Sovuch2Cn(ul)
                s3 = Sovuch3Cn(ul)
                s4 = Sovuch4Cn(ul)
                s5 = Sovuch5Cn(ul)
                if s1.get_pr(num):
                    num = s1.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s2.get_pr(num):
                    num = s2.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s3.get_pr(num):
                    num = s3.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s4.get_pr(num):
                    num = s4.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
                elif s5.get_pr(num):
                    num = s5.get()
                    rayon = rayon.get_rayon()
                    final.append([num, rayon])
                    kolvo = 1
        else:
            if i[0][:3] == 'Ул.' and i[0][-2:] == 'Ул':
                number = 1
                ul = i[0][4:-3]
            elif i[0][:3] == 'Ул.' or i[0][:3] == 'ул.':
                number = 1
                ul = i[0][4:]
            elif i[0][-2:] == 'Ул':
                number = 1
                ul = i[0][:-3]
            elif i[0][:4] == 'Пр-Т' or i[0][:4] == 'пр-т':
                number = 1
                ul = i[0][5:]
            elif i[0][-3:] == 'Б-Р':
                number = 1
                if i[0][:-4] != 'Победы':
                    ul = i[0][:-4]
                else:
                    ul = i[0]
            elif i[0][:4] == 'Пер.' and i[0][-3:] == 'Пер':
                ul = i[0][5:-4]
                number = 2
            elif i[0][:4] == 'Пер.' or i[0][:4] == 'пер.':
                ul = i[0][5:]
                number = 2
            elif i[0][-3:] == 'Пер':
                ul = i[0][:-4]
                number = 2
            else:
                ul = i[0]
                number = 3
            s1 = Sovuch1S(ul)
            s2 = Sovuch2S(ul)
            s3 = Sovuch3S(ul)
            s4 = Sovuch4S(ul)
            s5 = Sovuch5S(ul)
            s6 = Sovuch6S(ul)
            s7 = Sovuch7S(ul)
            s8 = Sovuch8S(ul)
            rayon = Sovetskiy(ul)
            if s1.get_sovuch(number):
                num = s1.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s2.get_sovuch(number):
                num = s2.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s3.get_sovuch(number):
                num = s3.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s4.get_sovuch(number):
                num = s4.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s5.get_sovuch(number):
                num = s5.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s6.get_sovuch(number):
                num = s6.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s7.get_sovuch(number):
                num = s7.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s8.get_sovuch(number):
                num = s8.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            s1 = Sovuch1Z(ul)
            s2 = Sovuch2Z(ul)
            s3 = Sovuch3Z(ul)
            s4 = Sovuch4Z(ul)
            s5 = Sovuch5Z(ul)
            rayon = Zheleznodorozhny(ul)
            if s1.get_sovuch(number):
                num = s1.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s2.get_sovuch(number):
                num = s2.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s3.get_sovuch(number):
                num = s3.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s4.get_sovuch(number):
                num = s4.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s5.get_sovuch(number):
                num = s5.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            s1 = Sovuch1K(ul)
            s2 = Sovuch2K(ul)
            s3 = Sovuch3K(ul)
            s4 = Sovuch4K(ul)
            s5 = Sovuch5K(ul)
            s6 = Sovuch6K(ul)
            s7 = Sovuch7K(ul)
            s8 = Sovuch8K(ul)
            s9 = Sovuch9K(ul)
            s10 = Sovuch10K(ul)
            s11 = Sovuch11K(ul)
            s12 = Sovuch12K(ul)
            rayon = Kominternovsky(ul)
            if s1.get_sovuch(number):
                num = s1.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s2.get_sovuch(number):
                num = s2.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s3.get_sovuch(number):
                num = s3.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s4.get_sovuch(number):
                num = s4.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s5.get_sovuch(number):
                num = s5.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s6.get_sovuch(number):
                num = s6.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s7.get_sovuch(number):
                num = s7.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s8.get_sovuch(number):
                num = s8.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s9.get_sovuch(number):
                num = s9.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s10.get_sovuch(number):
                num = s10.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s11.get_sovuch(number):
                num = s11.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s12.get_sovuch(number):
                num = s12.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            s1 = Sovuch1Lv(ul)
            s2 = Sovuch2Lv(ul)
            s3 = Sovuch3Lv(ul)
            s4 = Sovuch4Lv(ul)
            s5 = Sovuch5Lv(ul)
            s6 = Sovuch6Lv(ul)
            s7 = Sovuch7Lv(ul)
            s8 = Sovuch8Lv(ul)
            rayon = Levoberezhny(ul)
            if s1.get_sovuch(number):
                num = s1.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s2.get_sovuch(number):
                num = s2.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s3.get_sovuch(number):
                num = s3.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s4.get_sovuch(number):
                num = s4.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s5.get_sovuch(number):
                num = s5.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s6.get_sovuch(number):
                num = s6.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s7.get_sovuch(number):
                num = s7.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s8.get_sovuch(number):
                num = s8.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            s1 = Sovuch1Ln(ul)
            s2 = Sovuch2Ln(ul)
            s3 = Sovuch3Ln(ul)
            s4 = Sovuch4Ln(ul)
            s5 = Sovuch5Ln(ul)
            s6 = Sovuch6Ln(ul)
            s7 = Sovuch7Ln(ul)
            rayon = Leninsky(ul)
            if s1.get_sovuch(number):
                num = s1.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s2.get_sovuch(number):
                num = s2.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s3.get_sovuch(number):
                num = s3.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s4.get_sovuch(number):
                num = s4.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s5.get_sovuch(number):
                num = s5.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s6.get_sovuch(number):
                num = s6.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s7.get_sovuch(number):
                num = s7.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            s1 = Sovuch1Cn(ul)
            s2 = Sovuch2Cn(ul)
            s3 = Sovuch3Cn(ul)
            s4 = Sovuch4Cn(ul)
            s5 = Sovuch5Cn(ul)
            rayon = Сentralny(ul)
            if s1.get_sovuch(number):
                num = s1.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s2.get_sovuch(number):
                num = s2.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s3.get_sovuch(number):
                num = s3.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s4.get_sovuch(number):
                num = s4.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            elif s5.get_sovuch(number):
                num = s5.get()
                rayon = rayon.get_rayon()
                final.append([num, rayon])
                kolvo = 1
            if kolvo == 0:
                if '.' not in i[1].split()[1] and 'Г' not in i[1].split()[1] and 'В' not in i[1].split()[
                    1] and 'А' not in \
                        i[1].split()[1] and 'Б' not in \
                        i[1].split()[1] and '/' not in i[1].split()[1]:
                    s1 = Sovuch1S(ul)
                    s2 = Sovuch2S(ul)
                    s3 = Sovuch3S(ul)
                    s4 = Sovuch4S(ul)
                    s5 = Sovuch5S(ul)
                    s6 = Sovuch6S(ul)
                    s7 = Sovuch7S(ul)
                    s8 = Sovuch8S(ul)
                    rayon = Sovetskiy(ul)
                    if (ul == 'Ворошилова' and int((i[1].split())[1]) % 2 != 0 and (
                            19 >= int((i[1].split())[1]) >= 3)) or (
                            ul == 'Ворошилова' and int((i[1].split())[1]) % 2 == 0 and (
                            136 >= int((i[1].split())[1]) >= 16)) or (
                            ul == 'Пирогова' and int((i[1].split())[1]) % 2 != 0 and (
                            int((i[1].split())[1]) >= 23)) or (
                            ul == 'Пирогова' and int((i[1].split())[1]) % 2 == 0 and (int((i[1].split())[1]) >= 8)):
                        num = s1.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == '9 Января' and int((i[1].split())[1]) % 2 != 0 and (int((i[1].split())[1]) >= 81)) or (
                            ul == 'Торпедо' and int((i[1].split())[1]) % 2 == 0 and (
                            9 >= int((i[1].split())[1]) >= 1)):
                        num = s3.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Машиностроителей' and int((i[1].split())[1]) % 2 == 0 and (
                            94 >= int((i[1].split())[1]) >= 82)) or (
                            ul == 'Машиностроителей' and int((i[1].split())[1]) == 57) or (
                            ul == 'Свободы' and int((i[1].split())[1]) % 2 != 0 and (
                            89 >= int((i[1].split())[1]) >= 79)):
                        num = s5.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif ul == 'Бахметьева' and int((i[1].split())[1]) % 2 != 0:
                        num = s6.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Газовая' and int((i[1].split())[1]) % 2 != 0 and (
                            23 >= int((i[1].split())[1]) >= 1)) or (
                            ul == 'Газовая' and int((i[1].split())[1]) % 2 == 0 and (
                            16 >= int((i[1].split())[1]) >= 2)) or (
                            ul == 'Кривошеина' and int((i[1].split())[1]) % 2 == 0 and (
                            int((i[1].split())[1]) >= 66)):
                        num = s7.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Карла Либкнехта' and int((i[1].split())[1]) % 2 != 0 and (
                            57 >= int((i[1].split())[1]) >= 53)) or (
                            ul == 'Карла Либкнехта' and int((i[1].split())[1]) == 74) or (
                            ul == 'Карла Либкнехта' and int((i[1].split())[1]) == 76) or (
                            ul == 'Летчика Колесниченко' and int((i[1].split())[1]) % 2 == 0 and (
                            int((i[1].split())[1]) >= 42)) or (
                            ul == 'Моисеева' and int((i[1].split())[1]) % 2 == 0 and (
                            84 >= int((i[1].split())[1]) >= 40)):
                        num = s8.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    s1 = Sovuch1Z(ul)
                    s2 = Sovuch2Z(ul)
                    s3 = Sovuch3Z(ul)
                    s4 = Sovuch4Z(ul)
                    s5 = Sovuch5Z(ul)
                    rayon = Zheleznodorozhny(ul)
                    if ul == 'Ленинский Пр-т' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 119:
                        num = s1.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif ul == 'Ленинский Пр-т' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 124:
                        num = s2.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif ul == 'Димитрова' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) <= 95:
                        num = s3.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Минская' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) <= 43) or (
                            ul == 'Минская' and int((i[1].split())[1]) % 2 == 0):
                        num = s4.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Ильича' and 42 >= int((i[1].split())[1]) >= 1) or (
                            ul == 'Минская' and int((i[1].split())[1]) / 2 != 0 and int((i[1].split())[1]) >= 43) or (
                            ul == 'Старых Большевиков ' and int((i[1].split())[1]) <= 69) or (
                            ul == 'Ленинградская' and int((i[1].split())[1]) == 1):
                        num = s5.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    s1 = Sovuch1K(ul)
                    s2 = Sovuch2K(ul)
                    s3 = Sovuch3K(ul)
                    s4 = Sovuch4K(ul)
                    s5 = Sovuch5K(ul)
                    s6 = Sovuch6K(ul)
                    s7 = Sovuch7K(ul)
                    s8 = Sovuch8K(ul)
                    s9 = Sovuch9K(ul)
                    s10 = Sovuch10K(ul)
                    s11 = Sovuch11K(ul)
                    s12 = Sovuch12K(ul)
                    rayon = Kominternovsky(ul)
                    if ul == 'Московский' and int((i[1].split())[1]) % 2 != 0:
                        num = s1.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif ul == 'Генерала Лизюкова' and int((i[1].split())[1]) % 2 != 0:
                        num = s2.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == '9 Января' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 102) or (
                            ul == 'Краснодонская' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 21) or (
                            ul == 'Краснодонская' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 14):
                        num = s5.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Хользунова' and int((i[1].split())[1]) % 2 == 0) or (
                            ul == 'Хользунова' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 107):
                        num = s6.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif ul == 'Владимира Невского' and int((i[1].split())[1]) % 2 == 0:
                        num = s7.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Рабочий проспект' and int((i[1].split())[1]) % 2 != 0 and int(
                            (i[1].split())[1]) >= 47) or (
                            ul == 'Рабочий проспект' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 62) or (
                            ul == 'Транспортная' and int((i[1].split())[1]) % 2 != 0):
                        num = s8.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Машиностроителей' and int((i[1].split())[1]) % 2 == 0 and 80 >= int(
                            (i[1].split())[1]) >= 2) or (
                            ul == 'Машиностроителей' and int((i[1].split())[1]) % 2 != 0 and 51 >= int(
                        (i[1].split())[1]) >= 1):
                        num = s9.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif ul == 'Владимира Невского' and int((i[1].split())[1]) % 2 != 0:
                        num = s10.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Генерала Лизюкова' and int((i[1].split())[1]) % 2 == 0) or (
                            ul == 'Шишкова' and int((i[1].split())[1]) % 2 != 0 and 83 >= int((i[1].split())[1]) >= 1):
                        num = s11.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Московский' and int((i[1].split())[1]) % 2 == 0 and 116 >= int(
                            (i[1].split())[1]) >= 2) or (
                            ul == 'Газовая' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 18) or (
                            ul == 'Газовая' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 25) or (
                            ul == 'Хользунова' and int((i[1].split())[1]) % 2 != 0 and 105 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Шишкова' and int((i[1].split())[1]) % 2 == 0 and 136 >= int((i[1].split())[1]) >= 2):
                        num = s12.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    s1 = Sovuch1Lv(ul)
                    s2 = Sovuch2Lv(ul)
                    s3 = Sovuch3Lv(ul)
                    s4 = Sovuch4Lv(ul)
                    s5 = Sovuch5Lv(ul)
                    s6 = Sovuch6Lv(ul)
                    s7 = Sovuch7Lv(ul)
                    s8 = Sovuch8Lv(ul)
                    rayon = Levoberezhny(ul)
                    if ul == 'Новосибирская' and int((i[1].split())[1]) % 2 != 0:
                        num = s1.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif ul == 'Ростовская' and int((i[1].split())[1]) % 2 == 0:
                        num = s2.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Авиастроителей набережная' and 22 >= int((i[1].split())[1]) >= 1) or (
                            ul == 'Новосибирская' and int((i[1].split())[1]) % 2 == 0) or (
                            ul == 'Ростовская' and int((i[1].split())[1]) % 2 != 0):
                        num = s3.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Ильича' and int((i[1].split())[1]) >= 42) or (
                            ul == 'Ленинградская' and int((i[1].split())[1]) >= 2):
                        num = s4.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Ленинский' and int((i[1].split())[1]) % 2 == 0) or (
                            ul == 'Авиастроителей набережная' and int((i[1].split())[1]) >= 23):
                        num = s5.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Ленинский' and int((i[1].split())[1]) % 2 != 0 and 117 >= int((i[1].split())[1])) or (
                            ul == 'Димитрова' and int((i[1].split())[1]) % 2 == 0 and 102 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Старых Большевиков' and int((i[1].split())[1]) >= 83):
                        num = s6.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Димитрова' and int((i[1].split())[1]) >= 103) or (
                            ul == 'Павловский' and int((i[1].split())[1]) != 56):
                        num = s7.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    s1 = Sovuch1Ln(ul)
                    s2 = Sovuch2Ln(ul)
                    s3 = Sovuch3Ln(ul)
                    s4 = Sovuch4Ln(ul)
                    s5 = Sovuch5Ln(ul)
                    s6 = Sovuch6Ln(ul)
                    s7 = Sovuch7Ln(ul)
                    rayon = Leninsky(ul)
                    if (ul == '3-Го Интерн.' and int((i[1].split())[1]) % 2 != 0 and int(
                            (i[1].split())[1]) >= 35) or (
                            ul == '3-Го Интерн.' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 26) or (
                            ul == 'Бахметьева' and int((i[1].split())[1]) % 2 == 0 and 10 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Войкова' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 17) or (
                            ul == 'Войкова' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 22) or (
                            ul == 'Куколкина' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 31) or (
                            ul == 'Куколкина' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 28) or (
                            ul == 'Пирогова' and int((i[1].split())[1]) % 2 == 0 and 6 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Пирогова' and int((i[1].split())[1]) % 2 != 0 and 21 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Революции 1905 Г.' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 23) or (
                            ul == 'Революции 1905 Г.' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 62):
                        num = s1.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == '9 Января' and int((i[1].split())[1]) % 2 == 0 and 40 >= int(
                            (i[1].split())[1]) >= 2) or (
                            ul == '9 Января' and int((i[1].split())[1]) % 2 != 0 and 47 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Белостокская' and 14 >= int((i[1].split())[1]) >= 1) or (
                            ul == 'Белостокская' and int((i[1].split())[1]) >= 23) or (
                            ul == 'Бетховена' and int((i[1].split())[1]) % 2 == 0) or (
                            ul == 'Веры Фигнер' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 112) or (
                            ul == 'Веры Фигнер' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 91) or (
                            ul == 'Володарского' and 37 >= int((i[1].split())[1]) >= 1) or (
                            ul == 'Клубная' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 51) or (
                            ul == 'Клубная' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 52) or (
                            ul == 'Куколкина' and int((i[1].split())[1]) % 2 != 0 and 29 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Куколкина' and int((i[1].split())[1]) % 2 == 0 and 26 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Куцыгина' and int((i[1].split())[1]) >= 32) or (
                            ul == 'Никитинская' and int((i[1].split())[1]) % 2 != 0 and 49 >= int(
                        (i[1].split())[1]) >= 29) or (
                            ul == 'Орджоникидзе' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 35) or (
                            ul == 'Орджоникидзе' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 28) or (
                            ul == 'Свободы' and int((i[1].split())[1]) % 2 == 0 and 24 >= int((i[1].split())[1])) or (
                            ul == 'Свободы' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) <= 59) or (
                            ul == 'Софьи Перовской' and int((i[1].split())[1]) % 2 != 0 and 73 >= int(
                        (i[1].split())[1]) >= 67) or (
                            ul == 'Софьи Перовской' and int((i[1].split())[1]) % 2 == 0 and 108 >= int(
                        (i[1].split())[1]) >= 100):
                        num = s2.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == '5 Декабря' and int((i[1].split())[1]) % 2 == 0 and 36 >= int(
                            (i[1].split())[1]) >= 2) or (
                            ul == '5 Декабря' and int((i[1].split())[1]) % 2 != 0 and 39 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Белостокская' and 22 >= int((i[1].split())[1]) >= 15) or (
                            ul == 'Бетховена' and int((i[1].split())[1]) % 2 != 0) or (
                            ul == 'Веры Фигнер' and int((i[1].split())[1]) % 2 == 0 and 110 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Веры Фигнер' and int((i[1].split())[1]) % 2 != 0 and 89 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Гродненская' and int((i[1].split())[1]) % 2 == 0 and 38 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Гродненская' and int((i[1].split())[1]) % 2 != 0 and 91 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Клубная' and int((i[1].split())[1]) % 2 == 0 and 50 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Клубная' and int((i[1].split())[1]) % 2 != 0 and 49 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Краснознамённая' and int((i[1].split())[1]) % 2 != 0 and 45 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Одоевского' and 13 >= int((i[1].split())[1]) >= 1) or (
                            ul == 'Ф.Энгельса' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 39) or (
                            ul == 'Ф.Энгельса' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 58) or (
                            ul == 'Каховский' and 43 >= int((i[1].split())[1]) >= 1):
                        num = s3.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Иртышская' and int((i[1].split())[1]) % 2 == 0) or (
                            ul == 'Крылова' and int((i[1].split())[1]) % 2 == 0 and 94 >= int(
                        (i[1].split())[1]) >= 28) or (
                            ul == 'Крылова' and 93 >= int((i[1].split())[1]) >= 27) or (
                            ul == 'Краснознамённая' and int((i[1].split())[1]) % 2 != 0 and 171 >= int(
                        (i[1].split())[1]) >= 99) or (
                            ul == 'Куцыгина' and 31 >= int((i[1].split())[1]) >= 1) or (
                            ul == 'Одоевского' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 28) or (
                            ul == 'Одоевского' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 15) or (
                            ul == 'Онежская' and int((i[1].split())[1]) % 2 != 0) or (
                            ul == 'Островского' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 26) or (
                            ul == 'Островского' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 27) or (
                            ul == 'Плехановская' and int((i[1].split())[1]) % 2 != 0) or (
                            ul == 'Успенского' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 28) or (
                            ul == 'Успенского' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 27) or (
                            ul == 'Херсонская' and int((i[1].split())[1]) % 2 == 0 and 62 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Чапаева' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 42) or (
                            ul == 'Чехова' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 22) or (
                            ul == 'Чехова' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 25):
                        num = s4.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == '5 Декабря' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 38) or (
                            ul == '5 Декабря' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 41) or (
                            ul == 'Ворошилова' and int((i[1].split())[1]) % 2 == 0 and 14 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Госпитальная' and int((i[1].split())[1]) % 2 == 0 and 76 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Госпитальная' and int((i[1].split())[1]) % 2 != 0 and 57 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Гродненская' and int((i[1].split())[1]) % 2 == 0 and 88 >= int(
                        (i[1].split())[1]) >= 40) or (
                            ul == 'Карла Либкнехта' and int((i[1].split())[1]) % 2 == 0 and 72 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Карла Либкнехта' and int((i[1].split())[1]) % 2 != 0 and 51 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Карельская' and int((i[1].split())[1]) % 2 == 0 and 58 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Карельская' and int((i[1].split())[1]) % 2 != 0 and 65 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Краснознамённая' and int((i[1].split())[1]) % 2 == 0 and 158 >= int(
                        (i[1].split())[1]) >= 12) or (
                            ul == 'Краснознамённая' and int((i[1].split())[1]) % 2 != 0 and 97 >= int(
                        (i[1].split())[1]) >= 47) or (
                            ul == 'Колесниченко' and int((i[1].split())[1]) % 2 != 0) or (
                            ul == 'Колесниченко' and int((i[1].split())[1]) % 2 == 0 and 40 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Крылова' and int((i[1].split())[1]) % 2 == 0 and 26 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Крылова' and int((i[1].split())[1]) % 2 != 0 and 25 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Моисеева' and int((i[1].split())[1]) % 2 != 0) or (
                            ul == 'Моисеева' and int((i[1].split())[1]) % 2 == 0 and 40 >= int(
                        (i[1].split())[1]) >= 84) or (
                            ul == 'Одоевского' and int((i[1].split())[1]) % 2 == 0 and 26 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Островского' and int((i[1].split())[1]) % 2 == 0 and 24 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Островского' and int((i[1].split())[1]) % 2 != 0 and 25 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Петрозаводская' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 13) or (
                            ul == 'Петрозаводская' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 14) or (
                            ul == 'Успенского' and int((i[1].split())[1]) % 2 == 0 and 26 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Успенского' and int((i[1].split())[1]) % 2 != 0 and 25 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Чапаева' and int((i[1].split())[1]) % 2 == 0 and 40 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Чапаева' and int((i[1].split())[1]) % 2 != 0 and 29 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Чехова' and int((i[1].split())[1]) % 2 == 0 and 20 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Чехова' and int((i[1].split())[1]) % 2 != 0 and 23 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Каховского' and int((i[1].split())[1]) >= 44) or (
                            ul == 'Летчиков' and int((i[1].split())[1]) % 2 == 0) or (
                            ul == 'Туркменский' and 28 >= int((i[1].split())[1]) >= 1):
                        num = s5.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == '9 Января' and int((i[1].split())[1]) % 2 != 0 and 79 >= int(
                            (i[1].split())[1]) >= 49) or (
                            ul == '9 Января' and int((i[1].split())[1]) % 2 == 0 and 100 >= int(
                        (i[1].split())[1]) >= 42) or (
                            ul == 'Ворошилова' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 21) or (
                            ul == 'Госпитальная' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 78) or (
                            ul == 'Госпитальная' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 59) or (
                            ul == 'Иртышская' and int((i[1].split())[1]) % 2 != 0) or (
                            ul == 'Карельская' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 60) or (
                            ul == 'Карельская' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 67) or (
                            ul == 'Краснознамённая' and int((i[1].split())[1]) % 2 == 0 and 214 >= int(
                        (i[1].split())[1]) >= 160) or (
                            ul == 'Краснознамённая' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 173) or (
                            ul == 'Кривошеина' and int((i[1].split())[1]) % 2 != 0) or (
                            ul == 'Кривошеина' and int((i[1].split())[1]) % 2 == 0 and 64 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Львовская' and int((i[1].split())[1]) % 2 != 0 and 11 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Львовская' and int((i[1].split())[1]) % 2 == 0 and 12 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Матросова' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 8) or (
                            ul == 'Матросова' and int((i[1].split())[1]) % 2 != 0 and 207 >= int(
                        (i[1].split())[1]) >= 73) or (
                            ul == 'Онежская' and int((i[1].split())[1]) % 2 == 0) or (
                            ul == 'Острогожская' and int((i[1].split())[1]) % 2 == 0 and 158 >= int(
                        (i[1].split())[1])) or (
                            ul == 'Острогожская' and int((i[1].split())[1]) % 2 != 0 and 91 >= int(
                        (i[1].split())[1]) >= 57) or (
                            ul == 'Петрозаводская' and int((i[1].split())[1]) % 2 != 0 and 11 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Петрозаводская' and int((i[1].split())[1]) % 2 == 0 and 12 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Пушкинская' and 45 >= int((i[1].split())[1]) >= 8) or (
                            ul == 'Херсонская' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 64) or (
                            ul == 'Херсонская' and int((i[1].split())[1]) % 2 != 0) or (
                            ul == 'Моисеева' and int((i[1].split())[1]) % 2 == 0 and 40 >= int(
                        (i[1].split())[1]) >= 84) or (
                            ul == 'Туркменский' and int((i[1].split())[1]) >= 29):
                        num = s6.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Кольцовская' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 33) or (
                            ul == 'Кольцовская' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 52) or (
                            ul == 'Краснознамённая' and int((i[1].split())[1]) % 2 == 0 and 10 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Львовская' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 13) or (
                            ul == 'Львовская' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 14) or (
                            ul == 'Матросова' and int((i[1].split())[1]) % 2 != 0 and 73 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Матросова' and int((i[1].split())[1]) % 2 == 0 and 6 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Никитинская' and int((i[1].split())[1]) % 2 == 0 and 54 >= int(
                        (i[1].split())[1]) >= 36) or (
                            ul == 'Острогожская' and int((i[1].split())[1]) % 2 != 0 and 55 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Свободы' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 65) or (
                            ul == 'Свободы' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 26) or (
                            ul == 'Чапаева' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 31):
                        num = s7.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    s1 = Sovuch1Cn(ul)
                    s2 = Sovuch2Cn(ul)
                    s3 = Sovuch3Cn(ul)
                    s4 = Sovuch4Cn(ul)
                    s5 = Sovuch5Cn(ul)
                    rayon = Сentralny(ul)
                    if (ul == 'Плехановская' and int((i[1].split())[1]) % 2 == 0 and 38 >= int(
                            (i[1].split())[1]) >= 12) or (
                            ul == 'Рабочий проспект' and int((i[1].split())[1]) % 2 != 0 and 45 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Рабочий проспект' and int((i[1].split())[1]) % 2 == 0 and 60 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Шишкова' and int((i[1].split())[1]) % 2 != 0 and int((i[1].split())[1]) >= 85):
                        num = s1.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Берег реки' and 82 >= int((i[1].split())[1]) >= 47) or (
                            ul == 'Революции' and int((i[1].split()[1])) / 2 != 0 and 55 >= int(
                        (i[1].split())[1]) >= 21) or (
                            ul == 'Володарского' and int((i[1].split())[1]) % 2 == 0) or (
                            ul == 'Володарского' and int((i[1].split())[1]) % 2 != 0 and int(
                        (i[1].split())[1]) >= 38) or (
                            ul == 'Орджоникидзе' and 37 >= int((i[1].split())[1]) >= 1) or (
                            ul == 'Плехановская' and int((i[1].split())[1]) % 2 == 0 and 10 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Транспортная' and int((i[1].split())[1]) % 2 == 0):
                        num = s2.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Карла Маркса' and int((i[1].split())[1]) % 2 != 0 and int(
                            (i[1].split())[1]) >= 47) or (
                            ul == 'Карла Маркса' and int((i[1].split())[1]) >= 62) or (
                            ul == 'Пушкинская' and 7 >= int((i[1].split())[1]) >= 1) or (
                            ul == 'Ф.Энгельса' and int((i[1].split())[1]) % 2 != 0 and 37 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Ф.Энгельса' and int((i[1].split())[1]) % 2 == 0 and 56 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Шишкова' and int((i[1].split())[1]) % 2 == 0 and int((i[1].split())[1]) >= 140):
                        num = s3.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Войкова' and int((i[1].split())[1]) % 2 != 0 and 15 >= int(
                            (i[1].split())[1]) >= 1) or (
                            ul == 'Войкова' and int((i[1].split())[1]) % 2 == 0 and 20 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Кольцовская' and int((i[1].split())[1]) % 2 != 0 and 31 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Кольцовская' and int((i[1].split())[1]) % 2 == 0 and 46 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Никитинская' and int((i[1].split())[1]) % 2 != 0 and 27 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Никитинская' and int((i[1].split())[1]) % 2 == 0 and 34 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Революции 1905 Г.' and int((i[1].split())[1]) % 2 != 0 and 11 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Революции 1905 Г.' and int((i[1].split())[1]) % 2 == 0 and 60 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Республиканская' and int((i[1].split())[1]) % 2 != 0 and 25 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Республиканская' and int((i[1].split())[1]) % 2 == 0 and 30 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Свердлова' and 6 >= int((i[1].split())[1]) >= 5) or (
                            ul == 'Студенческая' and int((i[1].split())[1]) != 17) or (
                            ul == 'Урицкого' and 5 >= int((i[1].split())[1]) >= 1):
                        num = s4.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
                    elif (ul == 'Революции' and int((i[1].split())[1]) % 2 != 0 and 19 >= int(
                            (i[1].split())[1]) >= 1) or (
                            ul == 'Революции' and int((i[1].split())[1]) % 2 == 0 and 58 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Московский' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 118) or (
                            ul == 'Берег реки' and 50 >= int((i[1].split())[1]) >= 1) or (
                            ul == 'Карла Маркса' and int((i[1].split())[1]) % 2 != 0 and 45 >= int(
                        (i[1].split())[1]) >= 1) or (
                            ul == 'Карла Маркса' and int((i[1].split())[1]) % 2 == 0 and 60 >= int(
                        (i[1].split())[1]) >= 2) or (
                            ul == 'Плехановская' and int((i[1].split())[1]) % 2 == 0 and int(
                        (i[1].split())[1]) >= 40) or (
                            ul == '3-Го Интерн.' and int((i[1].split())[1]) % 2 != 0 and 33 >= int(
                        (i[1].split())[1])) or (
                            ul == '3-Го Интерн.' and int((i[1].split())[1]) % 2 == 0 and 24 >= int(
                        (i[1].split())[1])):
                        num = s5.get()
                        rayon = rayon.get_rayon()
                        final.append([num, rayon])
                        kolvo = 1
            if kolvo == 0:
                final.append('')
    schetchik = -1
    for i in final:
        schetchik += 1
        if i:
            ws['F' + str(7 + schetchik)] = 'Судебный участок №' + str(i[0]) + ' мирового судьи ' + i[
                1] + ' района г.Воронежа'
    wb2.save(file_name)


main()
